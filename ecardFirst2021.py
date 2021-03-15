#!/usr/bin/python  

import sys
import subprocess
import importlib.util

packages = ['openpyxl']
for package_name in packages:
    spec = importlib.util.find_spec(package_name)
    if spec is None:
        print("Installing python packages. Please wait...")
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', package_name]) 

import openpyxl
import os


pathCards = "1st-grading-cards"
infosExcel = "ecard-infos.xlsx"
cardTemplate = "ecard-template.xlsx"
infosSheets = ["Infos-Card-Male", "Infos-Card-Female", "1st-Summary-Male", "1st-Summary-Female"]


# always keep same length
selectedEndCol = [12, 21]
pasteStartRow = [2, 2]
templateSheets = ["Infos", "Grades"]
pasteSheet = ["infoSheet", "gradeSheet"]


def loadSheets(infosExcel, infosSheets):
    sheetsCard = []
    print("Loading", infosExcel)
    wb = openpyxl.load_workbook(infosExcel, data_only=True)
    for i in range(0, len(infosSheets)):
        sheetsCard.insert(i, "Sheet" + str(i))
        sheetsCard[i] = wb[infosSheets[i]]
    return sheetsCard


def makeCard(pathCards, sheetsCard, cardTemplate, infosSheets, selectedEndCol, pasteStartRow, templateSheets, pasteSheet):
    os.makedirs(pathCards, exist_ok=True)
    addSheetNum = [*range(0, len(infosSheets) - 1, 2)]
    for i in range(0, 2):
        for k in range(0, sheetsCard[i].max_row):
            studentRow = k + 2
            studentName = sheetsCard[i].cell(row=studentRow, column=2).value
            if not studentName:
                continue
            studentCode = sheetsCard[i].cell(row=studentRow, column=1).value
            fileName = studentCode + "-" + studentName + ".xlsx"
            print("Creating file:", fileName)
            template = openpyxl.load_workbook(cardTemplate) 
   
            for h in range(0, len(templateSheets)):
                pasteSheet[h] = template[templateSheets[h]]
                        
            for j in range(0, len(pasteStartRow)):
                selectedRow = copyRange(1, studentRow, selectedEndCol[j], studentRow, sheetsCard[i + addSheetNum[j]]) 
                pasteRow = pasteRange(1, pasteStartRow[j], selectedEndCol[j], pasteStartRow[j], pasteSheet[j], selectedRow)

            template.save(os.path.join(pathCards,fileName))
    print("The outputs are in the folder:", pathCards)
    

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


if __name__ == '__main__':
    sheetsCard = loadSheets(infosExcel, infosSheets)
    makeCard(pathCards, sheetsCard, cardTemplate, infosSheets, selectedEndCol, pasteStartRow, templateSheets, pasteSheet)
    print("Done!")